import os
import sys

sys.path.append('/home/dylan/catkin_ws/src/yolo_ros/src/')

import json
import numpy as np
import json
import skimage.io as io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from tensorflow.python.client import device_lib

import tensorflow as tf
from TensorFlow_Yolo.yolov3.utils import image_preprocess, postprocess_boxes, nms, draw_bbox, Load_Yolo_model, Create_Yolo
from TensorFlow_Yolo.yolov3.configs import *
import cv2

import matplotlib.pyplot as plt

# These are the imports from other files
from evaluate import *

print(device_lib.list_local_devices())

# You need to have bboxResults initialised as a dictionary of numpy arrays.
# bboxCFList is a list of numpy arrays in the same order as bboxResults

# Format your bbox and confidence level results (if needed) and append them in these variables
real_test_dir = '/home/dylan/catkin_ws/src/yolo_ros/src/TensorFlow_Yolo/IMAGES/image_30m/'
export_path = "/home/dylan/catkin_ws/src/yolo_ros/src/TensorFlow_Yolo/exports/results_30m.xlsx"
labelFilePath = "/home/dylan/catkin_ws/src/yolo_ros/src/TensorFlow_Yolo/labels/label_30m_2.json"

image_paths = []
for filename in sorted(os.listdir(real_test_dir)):
    print(filename)
    if os.path.splitext(filename)[1].lower() in ['.png', '.jpg', '.jpeg']:
        image_paths.append(os.path.join(real_test_dir, filename))
        
yolo=Load_Yolo_model()
input_size = YOLO_INPUT_SIZE
score_threshold=0.3
iou_threshold=0.45

bboxResults = {}
bboxCFList = []
        
for image_path in image_paths:
    img = io.imread(image_path)
    img_arr = np.array(img)

    image_data = image_preprocess(np.copy(img_arr), [input_size, input_size])
    image_data = image_data[np.newaxis, ...].astype(np.float32)

    batched_input = tf.constant(image_data)
    pred_bbox = yolo(batched_input)

    pred_bbox = [tf.reshape(x, (-1, tf.shape(x)[-1])) for x in pred_bbox]
    pred_bbox = tf.concat(pred_bbox, axis=0)

    bboxes = postprocess_boxes(pred_bbox, img_arr, input_size, score_threshold)
    bboxes = nms(bboxes, iou_threshold, method='nms')

    r = {"rois": [], "score": []}
    for i, bbox in enumerate(bboxes):
        coor = np.array(bbox[:4], dtype=np.int32)
        x1, y1, x2, y2 = coor[0], coor[1], coor[2], coor[3]
        score = bbox[4]
        r["rois"].append([y1, x1, y2, x2])
        r["score"].append(score)

    print(r)

    frame = draw_bbox(img_arr, bboxes, CLASSES=TRAIN_CLASSES, rectangle_colors=(255,0,0))

    bboxCFList.append(r["score"] )
    bboxResults[image_path.replace(real_test_dir, "")] = r['rois']
 
    
print(f"{image_paths}\n")
print(f"{bboxResults}\n")
print(f"{bboxCFList}\n")

# First convert dictionary of lists into a dataframe, with the image name as index
df = pd.DataFrame.from_dict(bboxResults,orient='index', columns=['BBox Array'])

# Convert list of Confidence Level numpy array into pandas Series and append into the dataframe
cfdf = pd.Series(bboxCFList)
df = df.assign(CF=cfdf.values)
print(df)


try: 
    df.to_excel(export_path, header=True)
except PermissionError:
    print("There is already an existing file. You should remove it if you want to overwrite the file.")


# First transfer raw bounding box values and confidence levels of each image into an excel sheet
transfer_bbox(export_path,labelFilePath)

# Set the column names of each column used
set_header(export_path)

# Calculate and store each evaluation matrix
totalImg = calculate_total_images(export_path, bboxResults)
fNegCount = calculate_false_negatives(export_path, bboxResults)
fPosCount = calculate_false_positives(export_path, bboxResults)
passingRate = calculate_passing_rate(export_path, bboxResults)
average_cf = calculate_avg_cf(export_path, bboxResults)

# Finally, clean up the excel sheet for easy reading
clean_excel(export_path)