import skimage
import pandas as pd
import numpy as np
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def set_header(outputFile):
    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active
    
    ws['A1'] = 'File Name'
    ws['A1'].font = Font(bold=True)
    
    ws['B1'] = 'BBox Array'
    ws['B1'].font = Font(bold=True)
    
    ws['C1'] = 'Confidence Level'
    ws['C1'].font = Font(bold=True)
    
    ws['D1'] = 'Remarks'
    ws['D1'].font = Font(bold=True)
    
    ws['F1'] = 'Evaluation Matrix'
    ws['F1'].font = Font(bold=True)
    ws['F1'].fill = PatternFill(fgColor="008000", fill_type = "solid")
    ws['F1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('F1:G3')
    
    ws['F4'] = 'Total Images'
    ws['F5'] = 'No. of FP'
    ws['F6'] = 'No. of FN'
    ws['F7'] = 'Avg CF Level'
    ws['F8'] = '% Pass'
    ws['G8'].number_format = '#,##0.00'
    
    ws['N1'] = 'Ground Truth'
    ws['N1'].font = Font(bold=True)
    ws['N1'].alignment = Alignment(horizontal="center", vertical="top")
    ws.merge_cells('N1:O50')
    
    ws['P1'] = 'BBox Array'
    ws['P1'].font = Font(bold=True)
    
    # This section here is to create the definition area
    ws['F12'] = 'Definition'
    ws['F12'].font = Font(bold=True)
    ws['F12'].fill = PatternFill(fgColor="FFFF00", fill_type = "solid")
    ws['F12'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('F12:L13')
    
    ws['F14'] = 'No. of FP'
    ws['G14'] = 'Number of false positives - Object detected is not what it is supposed to be'
    ws['G14'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('F14:F15')
    ws.merge_cells('G14:L15')
    
    ws['F16'] = 'No. of Misses'
    ws['G16'] = 'Number of Misses - Target object that should be detected by model, but was not detected'
    ws['G16'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('F16:F17')
    ws.merge_cells('G16:L17')
    
    ws['F18'] = 'Avg CF Level'
    ws['G18'] = 'Average Confidence Level - Average confidence level of images that passed (no FP or FN)' 
    ws['G18'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('F18:F19')
    ws.merge_cells('G18:L19')
    
    ws['F20'] = '% Pass'
    ws['G20'] = 'Percentage of images that passed the detection test with no false positives or false negatives'
    ws['G20'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('F20:F21')
    ws.merge_cells('G20:L21')
    
    
    # Update excel sheet
    wb.save(filename=outputFile)


def calculate_false_positives(outputFile, dict):

    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active

    rowNum = 2      # First entry is at row number 2
    fPosCount = 0

    # Pixel tolerance for false positive
    # If any of the values in the array exceeds +- ptol, flagged as false positive
    ptol = 100


    # Iterate through each np array and check for false positives
    for key in dict:

        # If there are more than 1, automatically false positive detected
        if len(dict[key]) > 4:
            ws.cell(row=rowNum, column=4, value='False Positive').font = Font(color='FF8C00', bold=True)
            fPosCount += 1

        # If there are only 1 object, compare it with ground truth value
        # First access the cell, then convert the string into numpy array
        elif len(dict[key]) == 4:
            gtString = ws.cell(row=rowNum, column=16).value
            gtString = gtString.replace('[', '')
            gtString = gtString.replace(']', '')

            gtArray = np.fromstring(gtString, dtype=int, sep=' ')

            # false indicates false positive (1 or more values exceeded tolerance)
            accurate = compareArrays(gtArray, dict[key][0].astype(float), ptol)


            if not accurate:
                ws.cell(row=rowNum, column=4, value='False Positive').font = Font(color='FF8C00', bold=True)
                fPosCount += 1

        rowNum += 1


    # Update excel sheet
    ws['G5'] = fPosCount
    wb.save(filename=outputFile)

    return fPosCount

def calculate_false_negatives(outputFile, dict):
    
    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active
    
    
    fNegCount = 0
    rowNum = 2     # First entry is at row number 2
    
    # Iterate through each np array and check for false negatives
    for key in dict:
        gtCell = ws.cell(row=rowNum, column=16).value
        # print(gtCell)
        if len(dict[key]) == 0 and gtCell != '[]':
            ws.cell(row=rowNum, column=4, value='False Negative').font = Font(color='FF0000', bold=True)
            fNegCount += 1
        rowNum += 1
    
    # Update excel sheet
    ws['G6'] = fNegCount
    wb.save(filename=outputFile)
    
    return fNegCount

def calculate_total_images(outputFile, dict):
    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active
    
    # Update excel sheet
    ws['G4'] = len(dict)
    wb.save(filename=outputFile)
    
    return len(dict)

def calculate_passing_rate(outputFile, dict):

    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active
    
    imageSize = calculate_total_images(outputFile, dict)
    passCount = 0
    
    for i in range(2, imageSize+2): # First entry is at row 2
        remarks = ws.cell(row=i, column=4).value
        if remarks != 'False Positive' and remarks != 'False Negative':
            ws.cell(row=i, column=4).value = 'Passed'
            ws.cell(row=i, column=4).font = Font(color='008000', bold=True)
            passCount += 1
            
    passing_rate = round(passCount/imageSize * 100, 2)
            
    
    # Update excel sheet
    ws['G8'] = passing_rate
    wb.save(filename=outputFile)
    
    return passing_rate

def calculate_avg_cf(outputFile, dict):
    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active
    
    imageSize = calculate_total_images(outputFile, dict)
    passCount = 0
    cfList = []
    
    for i in range(2, imageSize+2): # First entry is at row 2
        remarks = ws.cell(row=i, column=4).value
        if remarks == 'Passed':
            passCount += 1
            cfString = ws.cell(row=i, column=3).value
            cfString = cfString.replace('[', '')
            cfString = cfString.replace(']', '')
            cfList.append(float(cfString))
    
    avg_cf = round(sum(cfList) / passCount, 5)
    
    # Update excel sheet
    ws['G7'] = avg_cf
    wb.save(filename=outputFile)
    
    #return avg_cf

def clean_excel(outputFile):
    
    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active
    
    # Adjust column width
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value and cell.column_letter != "G": 
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 2
        
    # Update excel sheet
    wb.save(filename=outputFile)
    
    return


def transfer_bbox(outputFile, labelFilePath):
    # Load excel sheet for manipulation
    wb = load_workbook(filename=outputFile)
    ws = wb.active

    rowNum = 2  # First entry is at row number 2

    # Load json file
    with open(labelFilePath) as lf:
        lfData = json.load(lf)

    # Transfer bounding box values row by row
    for image in lfData["annotations"]:

        # Ground Truth bounding box list is in different format than mask RCNN rois
        gtList = [round(num) for num in image["bbox"]]

        # Reformats the list following mask RCNN rois for easier comparison
        if (len(gtList) == 4):
            reformattedList = [gtList[1], gtList[0], gtList[1]+gtList[3], gtList[0]+gtList[2]]

            # Check for negative numbers and bump them up to 0
            for i in range(len(reformattedList)):
                if reformattedList[i] < 0:
                    reformattedList[i] = 0

            # Afterwards convert to numpy array
            npArray = np.asarray(reformattedList)

        ws.cell(row=rowNum, column=16, value=str(npArray))
        rowNum += 1

    # Update excel sheet
    wb.save(filename=outputFile)

    return


def compareArrays(array1, array2, ptol):
    index = 0
    for i in array1:
        if (array1[index] - array2[index]) > ptol:
            return False
        elif (array2[index] - array1[index]) > ptol:
            return False
        index += 1
    return True